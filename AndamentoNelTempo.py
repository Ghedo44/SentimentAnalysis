# Salvo i dati su un excel
from openpyxl import Workbook, load_workbook
h=[now[:10], pos, neg, pos+neg, pos/(pos+neg)]
print(h)
# re-open and append
try:
  wb = load_workbook('/content/drive/MyDrive/SentimentAnalisys/excels/' + keyword + '_SentimentOverTime.xlsx')
  ws = wb.active
  #ws.delete_rows(idx=5)
  ws.append([now[:10], pos, neg, pos+neg, pos/(pos+neg)])
  wb.save('/content/drive/MyDrive/SentimentAnalisys/excels/' + keyword + '_SentimentOverTime.xlsx')
except:
  # create the file
  wb = Workbook()
  ws = wb.active
  ws.append(['Date', 'Pos', 'Neg', 'Tot', 'Perc'])
  ws.append([now[:10], pos, neg, pos+neg, pos/(pos+neg)])
  wb.save('/content/drive/MyDrive/SentimentAnalisys/excels/' + keyword + '_SentimentOverTime.xlsx')
  
  # Con i dati salvati sull'excel creo un grafico del sentimento 
xls_file = pd.ExcelFile('/content/drive/MyDrive/SentimentAnalisys/excels/' + keyword + '_SentimentOverTime.xlsx')

df4 = xls_file.parse()
print(df4.head())

import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.ticker as mtick
import numpy as np

df4['Date'] = df4['Date'].astype(str)
df4['Perc'] = df4['Perc']*100

# Some example data to display
fig, (ax1, ax2) = plt.subplots(2)
ax1.plot(df4['Date'], df4['Perc'], color = "g", marker = "o", linestyle= "--", linewidth = 3, markersize = 10)
ax1.yaxis.set_major_formatter(mtick.PercentFormatter())
ax1.set_title('% di commenti positivi')
ax2.plot(df4['Date'], df4['Tot'], color = "g", marker = "o", linestyle= "--", linewidth = 3, markersize = 10)
ax2.set_title('Numero di commenti')
fig.autofmt_xdate()
plt.savefig('/content/drive/MyDrive/SentimentAnalisys/charts/Subreddit=' + source + '_Keyword=' + keyword + '_SentimentOvertTime' + '.png')
plt.show()
